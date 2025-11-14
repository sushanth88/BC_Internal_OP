import os
import re
from datetime import date, datetime

from flask import Flask, render_template, redirect, url_for, flash, request, abort, make_response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from flask_wtf import FlaskForm, CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import (create_engine, Column, Integer, String, Date, Float, Text, Boolean, DateTime, ForeignKey, UniqueConstraint, func)
from sqlalchemy.orm import declarative_base, sessionmaker, relationship, scoped_session
from sqlalchemy import text
from wtforms import StringField, PasswordField, SubmitField, BooleanField, FloatField, IntegerField, DateField, TextAreaField
from flask_wtf.file import FileField as UploadFileField, FileAllowed, FileRequired
from wtforms.validators import DataRequired, Length
from sqlalchemy.exc import IntegrityError
import tempfile
import uuid
import shutil
import csv
import tempfile
import uuid
import shutil
from typing import Optional, List, Dict

BASE_DIR = os.path.dirname(__file__)
DB_PATH = os.path.join(BASE_DIR, 'data.sqlite')
SQLALCHEMY_DATABASE_URI = f'sqlite:///{DB_PATH}'

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('BC_SECRET', 'dev-secret')

# Detect optional Excel dependency once and reuse everywhere
try:
    import openpyxl as _openpyxl  # type: ignore
    OPENPYXL_AVAILABLE = True
    OPENPYXL_VERSION = getattr(_openpyxl, '__version__', 'unknown')
except Exception as _oxl_err:
    OPENPYXL_AVAILABLE = False
    OPENPYXL_VERSION = None
    OPENPYXL_IMPORT_ERROR = repr(_oxl_err)

# enable CSRF protection for forms and POST endpoints
csrf = CSRFProtect()
csrf.init_app(app)

engine = create_engine(SQLALCHEMY_DATABASE_URI, echo=False, future=True)
SessionLocal = scoped_session(sessionmaker(bind=engine))
Base = declarative_base()

login_manager = LoginManager(app)
login_manager.login_view = 'login'


class User(Base, UserMixin):
    __tablename__ = 'users'
    id = Column(Integer, primary_key=True)
    username = Column(String(50), unique=True, nullable=False)
    password_hash = Column(String(255), nullable=False)
    is_admin = Column(Boolean, default=False)
    created_at = Column(DateTime, default=datetime.utcnow)

    transactions = relationship('Transaction', back_populates='user')
    # when Transaction has multiple FKs to users (user_id, last_edited_by),
    # explicitly bind this relationship to the creator FK
    transactions = relationship('Transaction', back_populates='user', foreign_keys='Transaction.user_id')

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    @classmethod
    def create(cls, username, password, is_admin=False):
        return cls(username=username, password_hash=generate_password_hash(password), is_admin=is_admin)


class Transaction(Base):
    __tablename__ = 'transactions'
    # enforce uniqueness per date (only one transaction row per date globally)
    __table_args__ = (UniqueConstraint('date', name='uix_date'),)
    id = Column(Integer, primary_key=True)
    user_id = Column(Integer, ForeignKey('users.id'), nullable=False)
    date = Column(Date, nullable=False)

    total_net_sale = Column(Float, default=0.0)
    # number of orders (display above total_net_sale in form)
    number_of_orders = Column(Integer, default=0)
    net_card_tips = Column(Float, default=0.0)
    voids_cash_sale = Column(Float, default=0.0)
    after_discount_cash = Column(Float, default=0.0)
    dining_cash_and_tips = Column(Float, default=0.0)
    # tips collected for dine-in separate from dine-in cash
    dine_in_tips = Column(Float, default=0.0)
    party_orders_cash = Column(Float, default=0.0)
    biryani_po = Column(Float, default=0.0)
    event_hall = Column(Float, default=0.0)
    paid_to = Column(Float, default=0.0)
    total_cash = Column(Float, default=0.0)
    grubhub = Column(Float, default=0.0)
    doordash = Column(Float, default=0.0)
    uber_eats = Column(Float, default=0.0)
    cancelled_orders = Column(Float, default=0.0)
    # free-form notes field
    gross_revenue = Column(Float, default=0.0)
    notes = Column(Text, default='')
    # admin-only financial fields
    staff_commission = Column(Float, default=0.0)
    toast_fees = Column(Float, default=0.0)
    doordash_fees = Column(Float, default=0.0)
    uber_eats_fees = Column(Float, default=0.0)
    grubhub_fees = Column(Float, default=0.0)
    net_revenue = Column(Float, default=0.0)
    cash_verified_by_owner = Column(String(200), default='')

    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # track who last edited this transaction (user id)
    last_edited_by = Column(Integer, ForeignKey('users.id'), nullable=True)

    last_edited_user = relationship('User', foreign_keys=[last_edited_by])

    # explicitly bind the creator relationship to the user_id FK
    user = relationship('User', back_populates='transactions', foreign_keys=[user_id])


class AuditLog(Base):
    __tablename__ = 'audit_logs'
    id = Column(Integer, primary_key=True)
    actor_id = Column(Integer, ForeignKey('users.id'), nullable=False)
    action = Column(String(50), nullable=False)  # create/edit/delete
    tx_id = Column(Integer, nullable=True)
    timestamp = Column(DateTime, default=datetime.utcnow)
    details = Column(Text, default='')

    # bind actor relationship to actor_id FK to avoid ambiguity
    actor = relationship('User', foreign_keys=[actor_id])


class RestaurantExpense(Base):
    __tablename__ = 'restaurant_expenses'
    id = Column(Integer, primary_key=True)
    month = Column(String(20), default='')
    date = Column(Date, nullable=False)
    category = Column(String(100), default='')
    sub_category = Column(String(200), default='')
    card = Column(String(100), default='')
    card_member = Column(String(100), default='')
    account_number = Column(String(100), default='')
    transaction_type = Column(String(100), default='')
    distribution = Column(String(100), default='')
    vendor = Column(String(200), default='')
    description = Column(Text, default='')
    amount = Column(Float, default=0.0)
    payment_method = Column(String(50), default='')
    invoice_number = Column(String(100), default='')
    notes = Column(Text, default='')
    created_at = Column(DateTime, default=datetime.utcnow)


class InventoryReceiptItem(Base):
    __tablename__ = 'inventory_receipt_items'
    id = Column(Integer, primary_key=True)
    upc = Column(String(64), index=True, default='')
    name = Column(String(200), default='')
    quantity = Column(Float, default=0.0)
    total_price = Column(Float, default=0.0)
    vendor = Column(String(120), default='')
    receipt_date = Column(Date, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)


@login_manager.user_loader
def load_user(user_id):
    db = SessionLocal()
    return db.get(User, int(user_id))


@app.context_processor
def inject_today_tx():
    # expose whether the current logged-in user already has a transaction for today
    try:
        if not current_user.is_authenticated:
            return dict(today_tx_exists=False, today_tx_id=None)
        db = SessionLocal()
        # since transactions are unique per date globally, check existence by date
        exists = db.query(Transaction).filter(Transaction.date==date.today()).first()
        return dict(today_tx_exists=bool(exists), today_tx_id=(exists.id if exists else None))
    except Exception:
        return dict(today_tx_exists=False, today_tx_id=None)


class LoginForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(min=3, max=50)])
    password = PasswordField('Password', validators=[DataRequired()])
    submit = SubmitField('Login')


class TransactionForm(FlaskForm):
    date = DateField('Date', default=date.today(), validators=[DataRequired()])
    number_of_orders = IntegerField('Number of Orders', default=0)
    total_net_sale = FloatField('Total Voids', default=0.0)
    net_card_tips = FloatField('Net Card Tips', default=0.0)
    voids_cash_sale = FloatField('Voids Cash Sale (After 5% Discount)', default=0.0)
    after_discount_cash = FloatField('Toast Card Sale', default=0.0)
    dining_cash_and_tips = FloatField('Dine-In Cash', default=0.0)
    dine_in_tips = FloatField('Dine-In Tips', default=0.0)
    party_orders_cash = FloatField('Catering Orders (Exclude Biryani Trays)', default=0.0)
    biryani_po = FloatField('Biryani Trays', default=0.0)
    event_hall = FloatField('Event Hall Rental & Food', default=0.0)
    paid_to = FloatField('Paid to Employee (Add Name in Notes)', default=0.0)
    total_cash = FloatField('Total Cash', default=0.0)
    grubhub = FloatField('Grubhub', default=0.0)
    doordash = FloatField('Doordash', default=0.0)
    uber_eats = FloatField('Uber Eats', default=0.0)
    cancelled_orders = FloatField('Cancelled Orders', default=0.0)
    gross_revenue = FloatField('Gross Revenue', default=0.0)
    # Admin-only fields
    staff_commission = FloatField('Staff Commission', default=0.0)
    toast_fees = FloatField('Toast Fees', default=0.0)
    doordash_fees = FloatField('Doordash Fees', default=0.0)
    uber_eats_fees = FloatField('Uber Eats Fees', default=0.0)
    grubhub_fees = FloatField('Grubhub Fees', default=0.0)
    net_revenue = FloatField('Net Revenue', default=0.0)
    cash_verified_by_owner = StringField('Cash Verified By Owner', default='')
    notes = TextAreaField('Notes', default='')
    submit = SubmitField('Save')


class UploadTransactionsForm(FlaskForm):
    file = UploadFileField('Upload CSV or Excel (.csv, .xlsx)', validators=[
        FileRequired(),
        FileAllowed(['csv', 'xlsx'], 'CSV or Excel files only')
    ])
    overwrite = BooleanField('Overwrite existing dates (upsert)')
    submit = SubmitField('Import')


class BulkDeleteForm(FlaskForm):
    submit = SubmitField('Delete Selected')


class DeleteAllForm(FlaskForm):
    confirm = StringField('Type DELETE ALL to confirm', default='')
    submit = SubmitField('Delete ALL')


class RestaurantExpenseForm(FlaskForm):
    date = DateField('Date', validators=[DataRequired()], default=date.today())
    category = StringField('Category', default='')
    sub_category = StringField('Sub Category', default='')
    month = StringField('Month', default='')
    card = StringField('Card', default='')
    card_member = StringField('Card Member', default='')
    account_number = StringField('Account #', default='')
    transaction_type = StringField('Transaction Type', default='')
    distribution = StringField('Distribution', default='')
    vendor = StringField('Vendor', default='')
    description = TextAreaField('Description', default='')
    amount = FloatField('Amount', default=0.0, validators=[DataRequired()])
    payment_method = StringField('Payment Method', default='')
    invoice_number = StringField('Invoice #', default='')
    notes = TextAreaField('Notes', default='')
    submit = SubmitField('Add Expense')


class RestaurantExpenseUploadForm(FlaskForm):
    file = UploadFileField('Upload CSV or Excel (.csv, .xlsx)', validators=[
        FileRequired(),
        FileAllowed(['csv', 'xlsx'], 'CSV or Excel files only')
    ])
    overwrite = BooleanField('Overwrite existing rows with same Invoice #')
    submit = SubmitField('Import')


class InventoryUploadForm(FlaskForm):
    file = UploadFileField('Upload Inventory CSV or Excel (.csv, .xlsx)', validators=[
        FileRequired(),
        FileAllowed(['csv', 'xlsx'], 'CSV or Excel files only')
    ])
    overwrite = BooleanField('Overwrite existing items for same UPC/date')
    submit = SubmitField('Import Inventory')


def _safe_str(val) -> str:
    try:
        if val is None:
            return ''
        return str(val).strip()
    except Exception:
        return ''


def init_db():
    Base.metadata.create_all(bind=engine)
    # Deduplicate existing transactions: keep the earliest created_at per date (global)
    with engine.begin() as conn:
        # Restaurant expenses migrations: add missing columns to align with template
        try:
            conn.execute(text("SELECT month FROM restaurant_expenses LIMIT 1"))
        except Exception:
            try:
                conn.execute(text("ALTER TABLE restaurant_expenses ADD COLUMN month VARCHAR(20) DEFAULT ''"))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT sub_category FROM restaurant_expenses LIMIT 1"))
        except Exception:
            try:
                conn.execute(text("ALTER TABLE restaurant_expenses ADD COLUMN sub_category VARCHAR(200) DEFAULT ''"))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT card FROM restaurant_expenses LIMIT 1"))
        except Exception:
            try:
                conn.execute(text("ALTER TABLE restaurant_expenses ADD COLUMN card VARCHAR(100) DEFAULT ''"))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT card_member FROM restaurant_expenses LIMIT 1"))
        except Exception:
            try:
                conn.execute(text("ALTER TABLE restaurant_expenses ADD COLUMN card_member VARCHAR(100) DEFAULT ''"))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT account_number FROM restaurant_expenses LIMIT 1"))
        except Exception:
            try:
                conn.execute(text("ALTER TABLE restaurant_expenses ADD COLUMN account_number VARCHAR(100) DEFAULT ''"))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT transaction_type FROM restaurant_expenses LIMIT 1"))
        except Exception:
            try:
                conn.execute(text("ALTER TABLE restaurant_expenses ADD COLUMN transaction_type VARCHAR(100) DEFAULT ''"))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT distribution FROM restaurant_expenses LIMIT 1"))
        except Exception:
            try:
                conn.execute(text("ALTER TABLE restaurant_expenses ADD COLUMN distribution VARCHAR(100) DEFAULT ''"))
            except Exception:
                pass
        # Add last_edited_by column if missing (SQLite allows simple ADD COLUMN)
        try:
            conn.execute(text("SELECT last_edited_by FROM transactions LIMIT 1"))
        except Exception:
            # column missing -> add it
            try:
                conn.execute(text('ALTER TABLE transactions ADD COLUMN last_edited_by INTEGER'))
            except Exception:
                pass
        # Add new columns if missing: number_of_orders, dine_in_tips, notes
        try:
            conn.execute(text("SELECT number_of_orders FROM transactions LIMIT 1"))
        except Exception:
            try:
                conn.execute(text('ALTER TABLE transactions ADD COLUMN number_of_orders INTEGER DEFAULT 0'))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT dine_in_tips FROM transactions LIMIT 1"))
        except Exception:
            try:
                conn.execute(text('ALTER TABLE transactions ADD COLUMN dine_in_tips REAL DEFAULT 0.0'))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT notes FROM transactions LIMIT 1"))
        except Exception:
            try:
                conn.execute(text("ALTER TABLE transactions ADD COLUMN notes TEXT DEFAULT ''"))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT gross_revenue FROM transactions LIMIT 1"))
        except Exception:
            try:
                conn.execute(text('ALTER TABLE transactions ADD COLUMN gross_revenue REAL DEFAULT 0.0'))
            except Exception:
                pass
        # admin-only fields migration
        try:
            conn.execute(text("SELECT staff_commission FROM transactions LIMIT 1"))
        except Exception:
            try:
                conn.execute(text('ALTER TABLE transactions ADD COLUMN staff_commission REAL DEFAULT 0.0'))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT toast_fees FROM transactions LIMIT 1"))
        except Exception:
            try:
                conn.execute(text('ALTER TABLE transactions ADD COLUMN toast_fees REAL DEFAULT 0.0'))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT doordash_fees FROM transactions LIMIT 1"))
        except Exception:
            try:
                conn.execute(text('ALTER TABLE transactions ADD COLUMN doordash_fees REAL DEFAULT 0.0'))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT uber_eats_fees FROM transactions LIMIT 1"))
        except Exception:
            try:
                conn.execute(text('ALTER TABLE transactions ADD COLUMN uber_eats_fees REAL DEFAULT 0.0'))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT grubhub_fees FROM transactions LIMIT 1"))
        except Exception:
            try:
                conn.execute(text('ALTER TABLE transactions ADD COLUMN grubhub_fees REAL DEFAULT 0.0'))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT net_revenue FROM transactions LIMIT 1"))
        except Exception:
            try:
                conn.execute(text('ALTER TABLE transactions ADD COLUMN net_revenue REAL DEFAULT 0.0'))
            except Exception:
                pass
        try:
            conn.execute(text("SELECT cash_verified_by_owner FROM transactions LIMIT 1"))
        except Exception:
            try:
                conn.execute(text("ALTER TABLE transactions ADD COLUMN cash_verified_by_owner VARCHAR(200) DEFAULT ''"))
            except Exception:
                pass
        # If there was a previous unique index on (user_id, date), drop it
        conn.execute(text('DROP INDEX IF EXISTS uix_user_date'))
        # find groups with duplicates by date
        dup_rows = conn.execute(text("SELECT date, COUNT(*) AS cnt FROM transactions GROUP BY date HAVING COUNT(*)>1")).fetchall()
        for row in dup_rows:
            d = row[0]
            # select ids ordered by created_at ascending
            id_rows = conn.execute(text("SELECT id FROM transactions WHERE date=:d ORDER BY created_at ASC"), {'d': d}).fetchall()
            ids = [r[0] for r in id_rows]
            # keep first, delete others
            if len(ids) > 1:
                del_ids = ids[1:]
                id_list_sql = ','.join(str(int(x)) for x in del_ids)
                if id_list_sql:
                    conn.execute(text(f"DELETE FROM transactions WHERE id IN ({id_list_sql})"))
        # Now create unique index on date (will succeed since duplicates removed)
        conn.execute(text('CREATE UNIQUE INDEX IF NOT EXISTS uix_date ON transactions(date)'))


# Run migrations at import time to ensure the schema includes newly added columns.
try:
    init_db()
except Exception:
    # ignore migration errors at import time; runtime queries will surface issues
    pass


@app.route('/')
@login_required
def index():
    # Admin-only dashboard; regular users are redirected to Sales
    if not current_user.is_admin:
        flash('You do not have access to Dashboard.', 'warning')
        return redirect(url_for('sales'))
    db = SessionLocal()
    if current_user.is_admin:
        txs = db.query(Transaction).order_by(Transaction.date.desc()).limit(100).all()
        # Admin-only: build aggregates for charts
        try:
            from datetime import timedelta
            today_d = date.today()
            # DAILY: last 14 days (inclusive)
            d_start = today_d - timedelta(days=13)
            rows_daily = (
                db.query(Transaction.date, func.sum(Transaction.net_revenue))
                .filter(Transaction.date >= d_start)
                .group_by(Transaction.date)
                .order_by(Transaction.date)
                .all()
            )
            daily_map = {r[0]: float(r[1] or 0.0) for r in rows_daily}
            daily_labels = []
            daily_values = []
            for i in range(14):
                d = d_start + timedelta(days=i)
                daily_labels.append(d.strftime('%m-%d'))
                daily_values.append(round(daily_map.get(d, 0.0), 2))

            # WEEKLY: last 8 weeks grouped by week start (Monday)
            w_start = today_d - timedelta(weeks=7)
            rows_week = (
                db.query(Transaction.date, Transaction.net_revenue)
                .filter(Transaction.date >= w_start - timedelta(days=6))
                .all()
            )
            week_map = {}
            for d, val in rows_week:
                wk_start = d - timedelta(days=d.weekday())
                week_map[wk_start] = week_map.get(wk_start, 0.0) + float(val or 0.0)
            # build ordered last 8 week buckets
            weekly_labels = []
            weekly_values = []
            for i in range(8):
                wk = (w_start - timedelta(days=w_start.weekday())) + timedelta(weeks=i)
                weekly_labels.append(wk.strftime('Wk %W\n%b %d'))
                weekly_values.append(round(week_map.get(wk, 0.0), 2))

            # MONTHLY: last 12 months grouped by YYYY-MM
            m_start = (today_d.replace(day=1))
            # compute 12 months back
            from calendar import monthrange
            months = []
            y, m = m_start.year, m_start.month
            for _ in range(12):
                months.append((y, m))
                m -= 1
                if m == 0:
                    m = 12
                    y -= 1
            months = list(reversed(months))
            # query from first month start
            first_year, first_month = months[0]
            from_month_start = date(first_year, first_month, 1)
            rows_month = (
                db.query(func.strftime('%Y-%m', Transaction.date), func.sum(Transaction.net_revenue))
                .filter(Transaction.date >= from_month_start)
                .group_by(func.strftime('%Y-%m', Transaction.date))
                .order_by(func.strftime('%Y-%m', Transaction.date))
                .all()
            )
            month_map = {k: float(v or 0.0) for (k, v) in rows_month}
            monthly_labels = []
            monthly_values = []
            for (yy, mm) in months:
                key = f"{yy:04d}-{mm:02d}"
                monthly_labels.append(f"{yy}-{mm:02d}")
                monthly_values.append(round(month_map.get(key, 0.0), 2))
            # YEARLY: last 5 years grouped by year
            y_start = today_d.year - 4
            rows_year = (
                db.query(func.strftime('%Y', Transaction.date), func.sum(Transaction.net_revenue))
                .filter(func.strftime('%Y', Transaction.date) >= str(y_start))
                .group_by(func.strftime('%Y', Transaction.date))
                .order_by(func.strftime('%Y', Transaction.date))
                .all()
            )
            year_map = {k: float(v or 0.0) for (k, v) in rows_year}
            yearly_labels = []
            yearly_values = []
            for yy in range(y_start, today_d.year + 1):
                key = f"{yy:04d}"
                yearly_labels.append(key)
                yearly_values.append(round(year_map.get(key, 0.0), 2))
            # GROSS revenue series (daily/weekly/monthly/yearly)
            gross_daily_values = [0.0] * len(daily_values)
            for i in range(len(daily_labels)):
                # map/date lookup already computed in daily_map for net; recompute gross map
                pass
            rows_gross_daily = (
                db.query(Transaction.date, func.sum(Transaction.gross_revenue))
                .filter(Transaction.date >= d_start)
                .group_by(Transaction.date)
                .order_by(Transaction.date)
                .all()
            )
            gross_daily_map = {r[0]: float(r[1] or 0.0) for r in rows_gross_daily}
            gross_daily_values = [round(gross_daily_map.get(d_start + timedelta(days=i), 0.0), 2) for i in range(14)]

            # gross weekly
            gross_week_map = {}
            rows_gross_week = (
                db.query(Transaction.date, Transaction.gross_revenue)
                .filter(Transaction.date >= w_start - timedelta(days=6))
                .all()
            )
            for d, val in rows_gross_week:
                wk_start = d - timedelta(days=d.weekday())
                gross_week_map[wk_start] = gross_week_map.get(wk_start, 0.0) + float(val or 0.0)
            gross_weekly_values = [round(gross_week_map.get((w_start - timedelta(days=w_start.weekday())) + timedelta(weeks=i), 0.0), 2) for i in range(8)]

            # gross monthly
            rows_gross_month = (
                db.query(func.strftime('%Y-%m', Transaction.date), func.sum(Transaction.gross_revenue))
                .filter(Transaction.date >= from_month_start)
                .group_by(func.strftime('%Y-%m', Transaction.date))
                .order_by(func.strftime('%Y-%m', Transaction.date))
                .all()
            )
            gross_month_map = {k: float(v or 0.0) for (k, v) in rows_gross_month}
            gross_monthly_values = [round(gross_month_map.get(f"{yy:04d}-{mm:02d}", 0.0), 2) for (yy, mm) in months]

            # gross yearly
            rows_gross_year = (
                db.query(func.strftime('%Y', Transaction.date), func.sum(Transaction.gross_revenue))
                .filter(func.strftime('%Y', Transaction.date) >= str(y_start))
                .group_by(func.strftime('%Y', Transaction.date))
                .order_by(func.strftime('%Y', Transaction.date))
                .all()
            )
            gross_year_map = {k: float(v or 0.0) for (k, v) in rows_gross_year}
            gross_yearly_values = [round(gross_year_map.get(f"{yy:04d}", 0.0), 2) for yy in range(y_start, today_d.year + 1)]
        except Exception:
            # On any failure, gracefully degrade to empty series
            daily_labels, daily_values = [], []
            weekly_labels, weekly_values = [], []
            monthly_labels, monthly_values = [], []
        charts = {
            'daily': {'labels': daily_labels, 'values': daily_values},
            'weekly': {'labels': weekly_labels, 'values': weekly_values},
            'monthly': {'labels': monthly_labels, 'values': monthly_values},
            'yearly': {'labels': yearly_labels, 'values': yearly_values},
            'gross': {
                'daily': {'labels': daily_labels, 'values': gross_daily_values},
                'weekly': {'labels': weekly_labels, 'values': gross_weekly_values},
                'monthly': {'labels': monthly_labels, 'values': gross_monthly_values},
                'yearly': {'labels': yearly_labels, 'values': gross_yearly_values},
            },
        }
    else:
        # unreachable because non-admins are redirected above
        today = date.today()
        txs = db.query(Transaction).filter(Transaction.date == today).all()
        charts = None
    
    # pass today's date so template can decide whether regular users may edit
    return render_template('index.html', transactions=txs, today=date.today(), charts=charts)


@app.route('/sales')
@login_required
def sales():
    # Redirect Sales landing to the Transactions subpage for now
    return redirect(url_for('sales_transactions'))


@app.route('/sales/transactions')
@login_required
def sales_transactions():
    db = SessionLocal()
    # For admins show paginated transactions; regular users see today's transaction only.
    page = 1
    per_page = 50
    try:
        page = max(1, int(request.args.get('page', 1)))
    except Exception:
        page = 1
    try:
        per_page = max(1, min(200, int(request.args.get('per_page', 50))))
    except Exception:
        per_page = 50
    if current_user.is_admin:
        q = db.query(Transaction)
        total = q.count()
        txs = q.order_by(Transaction.date.desc()).offset((page-1)*per_page).limit(per_page).all()
        pages = (total + per_page - 1) // per_page if per_page else 1
        def make_url(p):
            return url_for('sales_transactions', page=p, per_page=per_page)
        pagination = {
            'page': page,
            'per_page': per_page,
            'total': total,
            'pages': pages,
            'has_prev': page > 1,
            'has_next': page < pages,
            'prev_url': make_url(page-1) if page > 1 else None,
            'next_url': make_url(page+1) if page < pages else None,
        }
    else:
        # regular users see only today's transaction (global)
        today_d = date.today()
        txs = db.query(Transaction).filter(Transaction.date == today_d).all()
        pagination = None
    bulk_delete_form = BulkDeleteForm()
    delete_all_form = DeleteAllForm()
    upload_form = UploadTransactionsForm()
    return render_template('sales_transactions.html', transactions=txs, today=date.today(), bulk_delete_form=bulk_delete_form, delete_all_form=delete_all_form, upload_form=upload_form, pagination=pagination)


@app.route('/catering')
@login_required
def catering():
    """Top-level Catering view: list transactions that have catering (party_orders_cash) > 0.
    Admins see recent entries; regular users see today's catering entries (if any).
    """
    db = SessionLocal()
    if current_user.is_admin:
        txs = db.query(Transaction).filter(Transaction.party_orders_cash > 0).order_by(Transaction.date.desc()).limit(200).all()
    else:
        today_d = date.today()
        txs = db.query(Transaction).filter(Transaction.date == today_d, Transaction.party_orders_cash > 0).all()
    return render_template('catering.html', transactions=txs, today=date.today())


# Expenses pages
@app.route('/expenses')
@login_required
def expenses():
    if not current_user.is_admin:
        flash('You do not have access to Expenses.', 'warning')
        return redirect(url_for('sales'))
    return render_template('expenses.html')


@app.route('/expenses/salaries')
@login_required
def expenses_salaries():
    if not current_user.is_admin:
        flash('You do not have access to Expenses.', 'warning')
        return redirect(url_for('sales'))
    return render_template('expenses_salaries.html')

@app.route('/expenses/inventory-tracking', methods=['GET', 'POST'])
@login_required
def expenses_inventory_tracking():
    if not current_user.is_admin:
        flash('You do not have access to Expenses.', 'warning')
        return redirect(url_for('sales'))
    form = InventoryUploadForm()
    db = SessionLocal()
    if request.method == 'POST' and form.validate_on_submit():
        f = form.file.data
        tmpdir = tempfile.mkdtemp(prefix='inv_upload_')
        path = os.path.join(tmpdir, f'upload_{uuid.uuid4().hex}')
        try:
            # Save uploaded file
            f.save(path)
            rows: List[Dict[str, str]] = []
            if path.lower().endswith('.csv'):
                with open(path, 'r', encoding='utf-8', errors='ignore') as fh:
                    reader = csv.DictReader(fh)
                    for r in reader:
                        rows.append(r)
            elif path.lower().endswith('.xlsx'):
                if not OPENPYXL_AVAILABLE:
                    flash('Excel support not available on server (openpyxl not installed). Please upload CSV.', 'warning')
                else:
                    from openpyxl import load_workbook  # type: ignore
                    wb = load_workbook(path, read_only=True)
                    ws = wb.active
                    headers = []
                    for j, cell in enumerate(ws[1], start=1):
                        headers.append(_safe_str(cell.value).lower())
                    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        rec = {}
                        for j, cell in enumerate(row):
                            key = headers[j] if j < len(headers) else f'col{j+1}'
                            rec[key] = _safe_str(cell.value)
                        rows.append(rec)
            else:
                flash('Unsupported file type', 'danger')

            # Heuristics for column names
            # Accept variations like UPC/upc, qty/quantity, total/price/amount, name/description, date
            def pick(d: dict, names: List[str]) -> str:
                for n in names:
                    if n in d:
                        return d.get(n)
                # fallback case-insensitive
                lower = {k.lower(): v for k, v in d.items()}
                for n in names:
                    if n.lower() in lower:
                        return lower[n.lower()]
                return ''

            added = 0
            updated = 0
            if rows:
                for r in rows:
                    upc = _safe_str(pick(r, ['upc','barcode','code']))
                    name = _safe_str(pick(r, ['name','item','description','product']))
                    qty = _parse_float(pick(r, ['qty','quantity','count','units']))
                    total = _parse_float(pick(r, ['total','total_price','price','amount','cost']))
                    vend = _safe_str(pick(r, ['vendor','supplier']))
                    dval = pick(r, ['date','receipt_date','purchased_at'])
                    rdate = None
                    try:
                        rdate = _parse_date(dval)
                    except Exception:
                        rdate = None

                    if not upc and not name:
                        continue

                    # overwrite behavior: if enabled, try to find same upc/date and replace
                    item = None
                    if form.overwrite.data and upc and rdate:
                        item = db.query(InventoryReceiptItem).filter(InventoryReceiptItem.upc==upc, InventoryReceiptItem.receipt_date==rdate).first()
                    if item:
                        item.name = name or item.name
                        item.quantity = float(qty or 0.0)
                        item.total_price = float(total or 0.0)
                        item.vendor = vend or item.vendor
                        updated += 1
                    else:
                        db.add(InventoryReceiptItem(upc=upc, name=name, quantity=float(qty or 0.0), total_price=float(total or 0.0), vendor=vend, receipt_date=rdate))
                        added += 1
                db.commit()
                if updated and not added:
                    flash(f'Updated {updated} items.', 'success')
                elif added and not updated:
                    flash(f'Imported {added} items.', 'success')
                else:
                    flash(f'Imported {added} items, updated {updated} items.', 'success')
            else:
                flash('No rows found in the uploaded file.', 'warning')
        except Exception as e:
            flash(f'Upload failed: {e}', 'danger')
        finally:
            try:
                shutil.rmtree(tmpdir, ignore_errors=True)
            except Exception:
                pass

    # Build aggregated view grouped by UPC
    rows = db.query(InventoryReceiptItem).all()
    aggregates: Dict[str, dict] = {}
    for it in rows:
        key = it.upc or (it.name or '').lower()
        if not key:
            key = f'unknown-{it.id}'
        agg = aggregates.get(key)
        if not agg:
            aggregates[key] = agg = dict(upc=it.upc, name=it.name, total_qty=0.0, total_spent=0.0, unit_price=0.0)
        agg['total_qty'] += (it.quantity or 0.0)
        agg['total_spent'] += (it.total_price or 0.0)
    for agg in aggregates.values():
        qty = agg['total_qty']
        agg['unit_price'] = (agg['total_spent'] / qty) if qty else 0.0

    # Sort by name then UPC for stable table
    aggregated_list = sorted(aggregates.values(), key=lambda a: (a.get('name') or '', a.get('upc') or ''))
    return render_template('expenses_inventory_tracking.html', form=form, items=aggregated_list)

 


@app.route('/expenses/restaurant', methods=['GET', 'POST'])
@login_required
def expenses_restaurant():
    if not current_user.is_admin:
        flash('You do not have access to Expenses.', 'warning')
        return redirect(url_for('sales'))
    db = SessionLocal()
    form = RestaurantExpenseForm()
    upload_form = RestaurantExpenseUploadForm()

    # Manual add
    if form.validate_on_submit() and request.method == 'POST' and request.form.get('form-id') == 'manual':
        exp = RestaurantExpense(
            month=(form.month.data or '').strip(),
            date=form.date.data,
            category=form.category.data.strip(),
            sub_category=(form.sub_category.data or '').strip(),
            card=(form.card.data or '').strip(),
            card_member=(form.card_member.data or '').strip(),
            account_number=(form.account_number.data or '').strip(),
            transaction_type=(form.transaction_type.data or '').strip(),
            distribution=(form.distribution.data or '').strip(),
            vendor=form.vendor.data.strip(),
            description=form.description.data.strip(),
            amount=float(form.amount.data or 0.0),
            payment_method=form.payment_method.data.strip(),
            invoice_number=form.invoice_number.data.strip(),
            notes=form.notes.data.strip(),
        )
        db.add(exp)
        db.commit()
        flash('Expense added', 'success')
        return redirect(url_for('expenses_restaurant'))

    # Filters (simple): by date and category
    q = db.query(RestaurantExpense)
    start = request.args.get('start')
    end = request.args.get('end')
    category = request.args.get('category', '').strip()
    if start:
        try:
            q = q.filter(RestaurantExpense.date >= _parse_date(start))
        except Exception:
            pass
    if end:
        try:
            q = q.filter(RestaurantExpense.date <= _parse_date(end))
        except Exception:
            pass
    if category:
        q = q.filter(RestaurantExpense.category.ilike(f"%{category}%"))
    # pagination
    try:
        page = max(1, int(request.args.get('page', 1)))
    except Exception:
        page = 1
    try:
        per_page = max(10, min(200, int(request.args.get('per_page', 50))))
    except Exception:
        per_page = 50
    total = q.count()
    pages = (total + per_page - 1) // per_page if per_page else 1
    rows = q.order_by(RestaurantExpense.date.desc(), RestaurantExpense.id.desc()).offset((page-1)*per_page).limit(per_page).all()
    # compute total amount across the full filtered set (not just current page)
    try:
        parsed_start = _parse_date(start) if start else None
    except Exception:
        parsed_start = None
    try:
        parsed_end = _parse_date(end) if end else None
    except Exception:
        parsed_end = None
    sum_q = db.query(func.sum(RestaurantExpense.amount))
    if parsed_start:
        sum_q = sum_q.filter(RestaurantExpense.date >= parsed_start)
    if parsed_end:
        sum_q = sum_q.filter(RestaurantExpense.date <= parsed_end)
    if category:
        sum_q = sum_q.filter(RestaurantExpense.category.ilike(f"%{category}%"))
    total_amount = sum_q.scalar() or 0.0
    def make_url(p):
        return url_for('expenses_restaurant', page=p, per_page=per_page, start=start or None, end=end or None, category=category or None)
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'pages': pages,
        'has_prev': page > 1,
        'has_next': page < pages,
        'prev_url': make_url(page-1) if page > 1 else None,
        'next_url': make_url(page+1) if page < pages else None,
    }

    return render_template('expenses_restaurant.html', form=form, upload_form=upload_form, expenses=rows, total_amount=total_amount, start=start or '', end=end or '', category=category, pagination=pagination)


@app.route('/expenses/restaurant/upload', methods=['POST'])
@login_required
def expenses_restaurant_upload():
    if not current_user.is_admin:
        flash('You do not have access to Expenses.', 'warning')
        return redirect(url_for('sales'))
    form = RestaurantExpenseUploadForm()
    if not form.validate_on_submit():
        for field, errs in form.errors.items():
            for e in errs:
                flash(f"{field}: {e}", 'danger')
        return redirect(url_for('expenses_restaurant'))

    f = form.file.data
    filename = getattr(f, 'filename', '') or ''
    name_lower = filename.lower()
    overwrite = bool(form.overwrite.data)
    tmpdir = tempfile.mkdtemp(prefix='exp_up_')
    path = os.path.join(tmpdir, filename or f'upload_{uuid.uuid4().hex}')
    f.save(path)

    added = 0
    updated = 0
    failed = 0
    errors = []
    db = SessionLocal()

    def normalize_header(h):
        return re.sub(r"[^a-z0-9]+", "_", (h or '').strip().lower())

    def map_row(row):
        # Accept a variety of column names, map to fields
        mapping = {
            'month': ['month'],
            'date': ['date', 'expense_date'],
            'card': ['card'],
            'card_member': ['card_member', 'cardmember', 'member'],
            'account_number': ['account_number', 'account_no', 'account', 'account_', 'account_number_'],
            'transaction_type': ['transaction_type', 'type'],
            'amount': ['amount', 'total', 'value'],
            'distribution': ['distribution'],
            'category': ['category'],
            'sub_category': ['sub_categories', 'subcategory', 'sub_category'],
            'vendor': ['vendor', 'payee', 'supplier'],
            'description': ['description', 'details', 'desc'],
            'payment_method': ['payment_method', 'method', 'pay_method'],
            'invoice_number': ['invoice_number', 'invoice', 'bill_no', 'invoice_no'],
            'notes': ['notes', 'note', 'remarks'],
        }
        out = {k: '' for k in mapping.keys()}
        for key, aliases in mapping.items():
            for a in aliases:
                if a in row and row[a] not in (None, ''):
                    out[key] = row[a]
                    break
        return out

    try:
        rows = []
        if name_lower.endswith('.xlsx'):
            if not OPENPYXL_AVAILABLE:
                app.logger.warning('openpyxl unavailable during Restaurant upload: %s', globals().get('OPENPYXL_IMPORT_ERROR', 'unknown error'))
                flash('Excel support requires openpyxl. Please install it or upload CSV.', 'danger')
                shutil.rmtree(tmpdir, ignore_errors=True)
                return redirect(url_for('expenses_restaurant'))
            wb = _openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            headers = [normalize_header((c.value or '')) for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for r in ws.iter_rows(min_row=2):
                row = {headers[i]: (r[i].value if i < len(r) else None) for i in range(len(headers))}
                rows.append(map_row(row))
        else:
            with open(path, 'r', newline='') as fh:
                reader = csv.DictReader(fh)
                headers = [normalize_header(h) for h in (reader.fieldnames or [])]
                for raw in reader:
                    row = {normalize_header(k): (v or '') for k, v in raw.items()}
                    rows.append(map_row(row))

        for r in rows:
            try:
                d = _parse_date(r.get('date')) if r.get('date') else None
                if not d:
                    failed += 1
                    errors.append('Missing date')
                    continue
                amt = float(r.get('amount') or 0.0)
                inv = str(r.get('invoice_number') or '').strip()
                if overwrite and inv:
                    existing = db.query(RestaurantExpense).filter(RestaurantExpense.invoice_number == inv).first()
                else:
                    existing = None
                if existing:
                    existing.date = d
                    existing.month = str(r.get('month') or '').strip()
                    existing.category = str(r.get('category') or '').strip()
                    existing.sub_category = str(r.get('sub_category') or '').strip()
                    existing.card = str(r.get('card') or '').strip()
                    existing.card_member = str(r.get('card_member') or '').strip()
                    existing.account_number = str(r.get('account_number') or '').strip()
                    existing.transaction_type = str(r.get('transaction_type') or '').strip()
                    existing.distribution = str(r.get('distribution') or '').strip()
                    existing.vendor = str(r.get('vendor') or '').strip()
                    existing.description = str(r.get('description') or '').strip()
                    existing.amount = amt
                    existing.payment_method = str(r.get('payment_method') or '').strip()
                    existing.notes = str(r.get('notes') or '').strip()
                    db.add(existing)
                    updated += 1
                else:
                    exp = RestaurantExpense(
                        month=str(r.get('month') or '').strip(),
                        date=d,
                        category=str(r.get('category') or '').strip(),
                        sub_category=str(r.get('sub_category') or '').strip(),
                        card=str(r.get('card') or '').strip(),
                        card_member=str(r.get('card_member') or '').strip(),
                        account_number=str(r.get('account_number') or '').strip(),
                        transaction_type=str(r.get('transaction_type') or '').strip(),
                        distribution=str(r.get('distribution') or '').strip(),
                        vendor=str(r.get('vendor') or '').strip(),
                        description=str(r.get('description') or '').strip(),
                        amount=amt,
                        payment_method=str(r.get('payment_method') or '').strip(),
                        invoice_number=inv,
                        notes=str(r.get('notes') or '').strip(),
                    )
                    db.add(exp)
                    added += 1
            except Exception as e:
                failed += 1
                errors.append(str(e))

        db.commit()
        msg = f"Imported {added} rows"
        if updated:
            msg += f", updated {updated}"
        if failed:
            msg += f", failed {failed}"
        flash(msg, 'success' if failed == 0 else 'warning')
    except Exception as e:
        app.logger.error('Restaurant expenses upload error: %s', e)
        flash(f'Upload error: {e}', 'danger')
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    return redirect(url_for('expenses_restaurant'))


@app.route('/expenses/restaurant/template.csv')
@login_required
def expenses_restaurant_template():
    if not current_user.is_admin:
        abort(403)
    headers = ['date', 'category', 'vendor', 'description', 'amount', 'payment_method', 'invoice_number', 'notes']
    content = ','.join(headers) + '\n'
    resp = make_response(content)
    resp.headers['Content-Type'] = 'text/csv'
    resp.headers['Content-Disposition'] = 'attachment; filename=restaurant_expenses_template.csv'
    return resp


@app.route('/expenses/guest-house')
@login_required
def expenses_guest_house():
    if not current_user.is_admin:
        flash('You do not have access to Expenses.', 'warning')
        return redirect(url_for('sales'))
    return render_template('expenses_guest_house.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        db = SessionLocal()
        user = db.query(User).filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data):
            login_user(user)
            flash('Logged in successfully.', 'success')
            return redirect(url_for('index'))
        flash('Invalid credentials', 'danger')
    return render_template('login.html', form=form)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out', 'info')
    return redirect(url_for('login'))


 


def can_edit_tx(tx: Transaction):
    # admin can edit any transaction. Regular users can edit only the transaction for today's date (global).
    if current_user.is_admin:
        return True
    return tx.date == date.today()


def _parse_date(val: str):
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    fmts = ['%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y']
    for f in fmts:
        try:
            return datetime.strptime(s, f).date()
        except Exception:
            pass
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    raise ValueError(f'Unrecognized date: {val!r}')


def _parse_float(val):
    try:
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().replace('\u00a0', ' ')
        if s == '':
            return 0.0
        for ch in ['$', ',', '€', '£']:
            s = s.replace(ch, '')
        return float(s)
    except Exception:
        return 0.0


def _row_to_tx_fields(row: dict):
    def _normalize_key(s: str) -> str:
        # normalize header: lowercase, spaces/dashes to underscore, remove punctuation, collapse repeats
        s = str(s).strip().lower()
        s = s.replace(' ', '_').replace('-', '_')
        s = re.sub(r'[^a-z0-9_]+', '', s)
        s = re.sub(r'_+', '_', s)
        return s

    # Build normalized mapping and apply synonyms for known long labels
    norm = {}
    synonyms = {
        # Long labels in the official template
        'catering_orders_exclude_biryani_trays': 'catering_orders',
        'event_hall_rental_food': 'event_hall',
        'paid_to_employee_add_name_in_notes': 'paid_to',
        # Common alternates that we already partially accept elsewhere
        'dine_in_cash': 'dine_in_cash',
        'biryani_trays': 'biryani_trays',
        'toast_card_sale': 'toast_card_sale',
        'total_voids': 'total_voids',
    }

    for k, v in row.items():
        key = _normalize_key(k)
        canon = synonyms.get(key, key)
        norm[canon] = v
    def get(*names):
        # exact match first
        for n in names:
            if n in norm:
                return norm[n]
        # fallback: try prefix/substring matches for robustness
        for n in names:
            for k in norm.keys():
                if k.startswith(n) or n in k:
                    return norm[k]
        return None
    fields = {}
    d = get('date')
    fields['date'] = _parse_date(d) if d is not None else None
    def f(*names):
        return _parse_float(get(*names))
    fields.update({
        'number_of_orders': int(f('number_of_orders') or 0),
        'total_net_sale': f('total_net_sale', 'total_voids', 'total'),
        'net_card_tips': f('net_card_tips'),
        'after_discount_cash': f('after_discount_cash', 'toast_card_sale'),
        'dining_cash_and_tips': f('dining_cash_and_tips', 'dine_in_cash'),
        'dine_in_tips': f('dine_in_tips'),
        'party_orders_cash': f('party_orders_cash', 'catering_orders', 'catering_orders_exclude_biryani_trays'),
        'biryani_po': f('biryani_po', 'biryani_trays'),
        'event_hall': f('event_hall', 'event_hall_rental_food'),
        'paid_to': f('paid_to', 'paid_to_employee_add_name_in_notes'),
        'grubhub': f('grubhub'),
        'doordash': f('doordash'),
        'uber_eats': f('uber_eats'),
        'cancelled_orders': f('cancelled_orders'),
        'toast_fees': f('toast_fees'),
        'doordash_fees': f('doordash_fees'),
        'uber_eats_fees': f('uber_eats_fees'),
        'grubhub_fees': f('grubhub_fees'),
        'notes': (get('notes') or '').strip() if isinstance(get('notes'), str) else (get('notes') or ''),
    })
    return fields


@app.route('/transactions/new', methods=['GET', 'POST'])
@login_required
def new_transaction():
    form = TransactionForm()
    db = SessionLocal()
    # If GET and today's transaction exists for this user, redirect to edit it.
    if request.method == 'GET':
        check_date = date.today()
        # since transactions are unique per date globally, check for any transaction for today
        exists_today = db.query(Transaction).filter(Transaction.date==check_date).first()
        if exists_today:
            flash('A transaction for today already exists. Redirecting to edit.', 'info')
            return redirect(url_for('edit_transaction', tx_id=exists_today.id))

    if form.validate_on_submit():
        # all users may only create for today
        if form.date.data != date.today():
            flash('Transactions may only be created for today.', 'warning')
            return redirect(url_for('new_transaction'))
        # enforce only one transaction per date (global)
        exists = db.query(Transaction).filter(Transaction.date==form.date.data).first()
        if exists:
            flash('A transaction for this date already exists. Edit the existing transaction instead.', 'warning')
            return redirect(url_for('edit_transaction', tx_id=exists.id))
        # compute derived values server-side
        paid_to_val = float(form.paid_to.data or 0.0)
        # Voids Cash Sale is computed as Total Voids * 0.95 (after 5% discount) and cannot be set by the client
        total_net_sale_val = float(form.total_net_sale.data or 0.0)
        voids = float(total_net_sale_val * 0.95)
        dineCash = float(form.dining_cash_and_tips.data or 0.0)
        dineTips = float(form.dine_in_tips.data or 0.0)
        party = float(form.party_orders_cash.data or 0.0)
        biryani = float(form.biryani_po.data or 0.0)
        eventHall = float(form.event_hall.data or 0.0)
        total_cash_val = voids + dineCash + dineTips + party + biryani + eventHall - paid_to_val

        afterDiscount = float(form.after_discount_cash.data or 0.0)
        netCard = float(form.net_card_tips.data or 0.0)
        doordash_val = float(form.doordash.data or 0.0)
        uber = float(form.uber_eats.data or 0.0)
        grub = float(form.grubhub.data or 0.0)
        cancelled = float(form.cancelled_orders.data or 0.0)
        gross_val = total_cash_val + (afterDiscount + netCard + doordash_val + uber + grub) - cancelled

        toastFeesVal = float(form.toast_fees.data or 0.0)
        doordashFeesVal = float(form.doordash_fees.data or 0.0)
        uberFeesVal = float(form.uber_eats_fees.data or 0.0)
        grubFeesVal = float(form.grubhub_fees.data or 0.0)
        net_val = gross_val - (toastFeesVal + doordashFeesVal + uberFeesVal + grubFeesVal)

        # compute staff commission as 10% of Catering Orders (party_orders_cash) and enforce server-side
        staff_commission_val = float((form.party_orders_cash.data or 0.0) * 0.10)

        tx = Transaction(
            user_id=current_user.id,
            date=form.date.data,
            total_net_sale=form.total_net_sale.data or 0.0,
            number_of_orders=form.number_of_orders.data or 0,
            net_card_tips=form.net_card_tips.data or 0.0,
            # enforce server-side voids value
            voids_cash_sale=voids,
            after_discount_cash=form.after_discount_cash.data or 0.0,
            dining_cash_and_tips=form.dining_cash_and_tips.data or 0.0,
            dine_in_tips=form.dine_in_tips.data or 0.0,
            party_orders_cash=form.party_orders_cash.data or 0.0,
            biryani_po=form.biryani_po.data or 0.0,
            event_hall=form.event_hall.data or 0.0,
            paid_to=paid_to_val,
            total_cash=total_cash_val,
            grubhub=form.grubhub.data or 0.0,
            doordash=form.doordash.data or 0.0,
            uber_eats=form.uber_eats.data or 0.0,
            cancelled_orders=form.cancelled_orders.data or 0.0,
            gross_revenue=gross_val,
            # admin-only financial fields: staff_commission is computed as 10% of party_orders_cash and cannot be posted
            staff_commission=staff_commission_val,
            toast_fees=(form.toast_fees.data or 0.0) if current_user.is_admin else 0.0,
            doordash_fees=(form.doordash_fees.data or 0.0) if current_user.is_admin else 0.0,
            uber_eats_fees=(form.uber_eats_fees.data or 0.0) if current_user.is_admin else 0.0,
            grubhub_fees=(form.grubhub_fees.data or 0.0) if current_user.is_admin else 0.0,
            # net_revenue is computed server-side and authoritative
            net_revenue=net_val,
            cash_verified_by_owner=(form.cash_verified_by_owner.data or '') if current_user.is_admin else '',
            notes=form.notes.data or '',
        )
        tx.last_edited_by = current_user.id
        db.add(tx)
        try:
            db.commit()
        except IntegrityError:
            db.rollback()
            # check by date globally
            existing = db.query(Transaction).filter(Transaction.date==tx.date).first()
            if existing:
                flash('A transaction for this date was just created. Redirecting to edit.', 'warning')
                return redirect(url_for('edit_transaction', tx_id=existing.id))
            flash('Could not save transaction due to a database error.', 'danger')
            return redirect(url_for('index'))
        # audit
        log = AuditLog(actor_id=current_user.id, action='create', tx_id=tx.id, details=f'Created transaction for {tx.date} by user {current_user.username}; last_edited_by={current_user.username}')
        db.add(log)
        db.commit()
        flash('Transaction saved', 'success')
        return redirect(url_for('index'))
    return render_template('transaction_form.html', form=form, action='New')


@app.route('/transactions/<int:tx_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_transaction(tx_id):
    db = SessionLocal()
    tx = db.get(Transaction, tx_id)
    if not tx:
        abort(404)
    if not can_edit_tx(tx):
        flash('You do not have permission to edit this transaction.', 'danger')
        return redirect(url_for('index'))
    form = TransactionForm(obj=tx)
    if form.validate_on_submit():
        if not current_user.is_admin and form.date.data != date.today():
            flash('Regular users may only set date to today.', 'warning')
            return redirect(url_for('edit_transaction', tx_id=tx_id))
        # Prevent changing the date to one that would duplicate another transaction for the same user
        db = SessionLocal()
        conflict = db.query(Transaction).filter(Transaction.user_id == tx.user_id, Transaction.date == form.date.data, Transaction.id != tx.id).first()
        if conflict:
            flash('Cannot change date — another transaction for this user already exists on that date.', 'warning')
            return redirect(url_for('edit_transaction', tx_id=tx_id))

        # update fields
        old = {f:getattr(tx,f) for f in ['date','number_of_orders','total_net_sale','net_card_tips','voids_cash_sale','after_discount_cash','dining_cash_and_tips','dine_in_tips','party_orders_cash','biryani_po','event_hall','paid_to','total_cash','grubhub','doordash','uber_eats','cancelled_orders','gross_revenue','staff_commission','toast_fees','doordash_fees','uber_eats_fees','grubhub_fees','net_revenue','cash_verified_by_owner','notes']}
        # set basic fields from form for non-computed and non-admin fields
        # We'll overwrite computed fields server-side below. Also, ignore admin-only fields for non-admin users.
        for field in ['date','number_of_orders','total_net_sale','net_card_tips','after_discount_cash','dining_cash_and_tips','dine_in_tips','party_orders_cash','biryani_po','event_hall','paid_to','grubhub','doordash','uber_eats','cancelled_orders','notes']:
            setattr(tx, field, getattr(form, field).data)
        # recompute derived values (server-side)
        paid_to_val = float(form.paid_to.data or 0.0)
        # Voids Cash Sale computed from total_net_sale
        total_net_sale_val = float(form.total_net_sale.data or 0.0)
        voids = float(total_net_sale_val * 0.95)
        dineCash = float(form.dining_cash_and_tips.data or 0.0)
        dineTips = float(form.dine_in_tips.data or 0.0)
        party = float(form.party_orders_cash.data or 0.0)
        biryani = float(form.biryani_po.data or 0.0)
        eventHall = float(form.event_hall.data or 0.0)
        total_cash_val = voids + dineCash + dineTips + party + biryani + eventHall - paid_to_val
        tx.total_cash = total_cash_val

        afterDiscount = float(form.after_discount_cash.data or 0.0)
        netCard = float(form.net_card_tips.data or 0.0)
        doordash_val = float(form.doordash.data or 0.0)
        uber = float(form.uber_eats.data or 0.0)
        grub = float(form.grubhub.data or 0.0)
        cancelled = float(form.cancelled_orders.data or 0.0)
        gross_val = total_cash_val + (afterDiscount + netCard + doordash_val + uber + grub) - cancelled
        tx.gross_revenue = gross_val

        toastFeesVal = float(form.toast_fees.data or 0.0)
        doordashFeesVal = float(form.doordash_fees.data or 0.0)
        uberFeesVal = float(form.uber_eats_fees.data or 0.0)
        grubFeesVal = float(form.grubhub_fees.data or 0.0)
        net_val = gross_val - (toastFeesVal + doordashFeesVal + uberFeesVal + grubFeesVal)
        # compute staff commission as 10% of Catering Orders (party_orders_cash) and enforce server-side
        staff_commission_val = float((form.party_orders_cash.data or 0.0) * 0.10)
        # admin-only fields: only overwrite with posted values if current_user is admin
        if current_user.is_admin:
            tx.staff_commission = staff_commission_val
            tx.toast_fees = form.toast_fees.data or 0.0
            tx.doordash_fees = form.doordash_fees.data or 0.0
            tx.uber_eats_fees = form.uber_eats_fees.data or 0.0
            tx.grubhub_fees = form.grubhub_fees.data or 0.0
            # net_revenue remains computed server-side for consistency
            tx.net_revenue = net_val
            tx.cash_verified_by_owner = form.cash_verified_by_owner.data or ''
        else:
            # ensure non-admins cannot set admin fields; staff_commission is still computed
            tx.staff_commission = staff_commission_val
            tx.toast_fees = 0.0
            tx.doordash_fees = 0.0
            tx.uber_eats_fees = 0.0
            tx.grubhub_fees = 0.0
            tx.net_revenue = net_val
            tx.cash_verified_by_owner = ''
        tx.last_edited_by = current_user.id
        db.add(tx)
        db.commit()
        # audit
        changes = []
        for k,v in old.items():
            if getattr(tx,k) != v:
                changes.append(f'{k}: {v} -> {getattr(tx,k)}')
        log = AuditLog(actor_id=current_user.id, action='edit', tx_id=tx.id, details='; '.join(changes) or 'no changes')
        db.add(log)
        db.commit()
        flash('Transaction updated', 'success')
        return redirect(url_for('index'))
    return render_template('transaction_form.html', form=form, action='Edit')


@app.route('/transactions/<int:tx_id>/delete', methods=['POST'])
@login_required
def delete_transaction(tx_id):
    db = SessionLocal()
    tx = db.get(Transaction, tx_id)
    if not tx:
        abort(404)
    if not can_edit_tx(tx):
        flash('You do not have permission to delete this transaction.', 'danger')
        return redirect(url_for('index'))
    db.delete(tx)
    db.commit()
    # audit
    log = AuditLog(actor_id=current_user.id, action='delete', tx_id=tx_id, details=f'Deleted transaction {tx_id}')
    db.add(log)
    db.commit()
    flash('Transaction deleted', 'info')
    return redirect(url_for('index'))


@app.route('/admin/transactions/delete-selected', methods=['POST'])
@login_required
def admin_delete_selected():
    if not current_user.is_admin:
        abort(403)
    form = BulkDeleteForm()
    if not form.validate_on_submit():
        flash('Invalid request.', 'danger')
        return redirect(url_for('sales'))
    ids = request.form.getlist('ids')
    try:
        ids = [int(x) for x in ids]
    except Exception:
        ids = []
    if not ids:
        flash('No transactions selected.', 'warning')
        return redirect(url_for('sales'))
    db = SessionLocal()
    deleted = 0
    for _id in ids:
        tx = db.get(Transaction, _id)
        if not tx:
            continue
        db.delete(tx)
        db.commit()
        log = AuditLog(actor_id=current_user.id, action='delete', tx_id=_id, details='[bulk] Deleted via Delete Selected')
        db.add(log)
        db.commit()
        deleted += 1
    flash(f'Deleted {deleted} transactions.', 'success')
    return redirect(url_for('sales'))


@app.route('/admin/transactions/delete-all', methods=['POST'])
@login_required
def admin_delete_all():
    if not current_user.is_admin:
        abort(403)
    form = DeleteAllForm()
    if not form.validate_on_submit():
        flash('Invalid request.', 'danger')
        return redirect(url_for('sales'))
    confirm_text = (form.confirm.data or '').strip()
    if confirm_text != 'DELETE ALL':
        flash("To delete ALL transactions, type 'DELETE ALL' in the confirmation box.", 'warning')
        return redirect(url_for('sales'))
    db = SessionLocal()
    count = db.query(Transaction).count()
    db.query(Transaction).delete()
    db.commit()
    log = AuditLog(actor_id=current_user.id, action='delete', tx_id=None, details=f'[bulk] Deleted ALL transactions (count={count})')
    db.add(log)
    db.commit()
    flash(f'All transactions deleted (count={count}).', 'success')
    return redirect(url_for('sales'))


@app.route('/admin/upload', methods=['GET', 'POST'])
@login_required
def admin_upload():
    if not current_user.is_admin:
        abort(403)
    form = UploadTransactionsForm()
    summary = None
    if form.validate_on_submit():
        file = form.file.data
        overwrite = bool(form.overwrite.data)
        filename = file.filename or ''
        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        rows = []
        errors = []
        try:
            if ext == 'csv':
                import io, csv as _csv
                content = file.read()
                try:
                    text = content.decode('utf-8-sig')
                except Exception:
                    text = content.decode('latin-1')
                reader = _csv.DictReader(io.StringIO(text))
                rows = list(reader)
            elif ext == 'xlsx':
                if not OPENPYXL_AVAILABLE:
                    app.logger.warning('openpyxl unavailable during Transactions upload: %s', globals().get('OPENPYXL_IMPORT_ERROR', 'unknown error'))
                    flash('Excel support requires openpyxl. Please install it or upload CSV.', 'danger')
                    return render_template('admin_upload.html', form=form)
                wb = _openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for r in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: r[i] for i in range(len(headers))})
            else:
                flash('Unsupported file type. Upload a .csv or .xlsx file.', 'danger')
                return render_template('admin_upload.html', form=form)
        except Exception as e:
            # Log full traceback for easier debugging in server logs
            try:
                import traceback
                app.logger.error('Upload read error: %s', e)
                traceback.print_exc()
            except Exception:
                pass
            flash(f'Failed to read file: {e}', 'danger')
            return render_template('admin_upload.html', form=form)

        db = SessionLocal()
        created = 0
        updated = 0
        skipped = 0
        seen_dates = set()
        for idx, row in enumerate(rows, start=2):
            try:
                data = _row_to_tx_fields(row)
                d = data.get('date')
                if not d:
                    skipped += 1
                    errors.append(f'Row {idx}: missing/invalid date')
                    continue
                if d in seen_dates:
                    skipped += 1
                    errors.append(f'Row {idx}: duplicate date in file {d}')
                    continue
                seen_dates.add(d)
                total_net_sale_val = float(data['total_net_sale'] or 0.0)
                voids = float(total_net_sale_val * 0.95)
                dineCash = float(data['dining_cash_and_tips'] or 0.0)
                dineTips = float(data['dine_in_tips'] or 0.0)
                party = float(data['party_orders_cash'] or 0.0)
                biryani = float(data['biryani_po'] or 0.0)
                eventHall = float(data['event_hall'] or 0.0)
                paid_to_val = float(data['paid_to'] or 0.0)
                total_cash_val = voids + dineCash + dineTips + party + biryani + eventHall - paid_to_val
                afterDiscount = float(data['after_discount_cash'] or 0.0)
                netCard = float(data['net_card_tips'] or 0.0)
                doordash_val = float(data['doordash'] or 0.0)
                uber = float(data['uber_eats'] or 0.0)
                grub = float(data['grubhub'] or 0.0)
                cancelled = float(data['cancelled_orders'] or 0.0)
                gross_val = total_cash_val + (afterDiscount + netCard + doordash_val + uber + grub) - cancelled
                toastFeesVal = float(data['toast_fees'] or 0.0)
                doordashFeesVal = float(data['doordash_fees'] or 0.0)
                uberFeesVal = float(data['uber_eats_fees'] or 0.0)
                grubFeesVal = float(data['grubhub_fees'] or 0.0)
                net_val = gross_val - (toastFeesVal + doordashFeesVal + uberFeesVal + grubFeesVal)
                staff_commission_val = float((data['party_orders_cash'] or 0.0) * 0.10)

                existing = db.query(Transaction).filter(Transaction.date == d).first()
                if existing:
                    if not overwrite:
                        skipped += 1
                        continue
                    old_vals = {k: getattr(existing, k) for k in ['date','number_of_orders','total_net_sale','net_card_tips','voids_cash_sale','after_discount_cash','dining_cash_and_tips','dine_in_tips','party_orders_cash','biryani_po','event_hall','paid_to','total_cash','grubhub','doordash','uber_eats','cancelled_orders','gross_revenue','staff_commission','toast_fees','doordash_fees','uber_eats_fees','grubhub_fees','net_revenue','cash_verified_by_owner','notes']}
                    for field in ['number_of_orders','total_net_sale','net_card_tips','after_discount_cash','dining_cash_and_tips','dine_in_tips','party_orders_cash','biryani_po','event_hall','paid_to','grubhub','doordash','uber_eats','cancelled_orders','notes']:
                        setattr(existing, field, data.get(field))
                    existing.voids_cash_sale = voids
                    existing.total_cash = total_cash_val
                    existing.gross_revenue = gross_val
                    existing.staff_commission = staff_commission_val
                    existing.toast_fees = toastFeesVal
                    existing.doordash_fees = doordashFeesVal
                    existing.uber_eats_fees = uberFeesVal
                    existing.grubhub_fees = grubFeesVal
                    existing.net_revenue = net_val
                    existing.last_edited_by = current_user.id
                    db.add(existing)
                    db.commit()
                    changes = []
                    for k, v in old_vals.items():
                        if getattr(existing, k) != v:
                            changes.append(f'{k}: {v} -> {getattr(existing,k)}')
                    log = AuditLog(actor_id=current_user.id, action='edit', tx_id=existing.id, details='[bulk upload] ' + ('; '.join(changes) or 'no changes'))
                    db.add(log)
                    db.commit()
                    updated += 1
                else:
                    tx = Transaction(
                        user_id=current_user.id,
                        date=d,
                        number_of_orders=data['number_of_orders'] or 0,
                        total_net_sale=data['total_net_sale'] or 0.0,
                        net_card_tips=data['net_card_tips'] or 0.0,
                        voids_cash_sale=voids,
                        after_discount_cash=data['after_discount_cash'] or 0.0,
                        dining_cash_and_tips=data['dining_cash_and_tips'] or 0.0,
                        dine_in_tips=data['dine_in_tips'] or 0.0,
                        party_orders_cash=data['party_orders_cash'] or 0.0,
                        biryani_po=data['biryani_po'] or 0.0,
                        event_hall=data['event_hall'] or 0.0,
                        paid_to=data['paid_to'] or 0.0,
                        total_cash=total_cash_val,
                        grubhub=data['grubhub'] or 0.0,
                        doordash=data['doordash'] or 0.0,
                        uber_eats=data['uber_eats'] or 0.0,
                        cancelled_orders=data['cancelled_orders'] or 0.0,
                        gross_revenue=gross_val,
                        staff_commission=staff_commission_val,
                        toast_fees=toastFeesVal,
                        doordash_fees=doordashFeesVal,
                        uber_eats_fees=uberFeesVal,
                        grubhub_fees=grubFeesVal,
                        net_revenue=net_val,
                        cash_verified_by_owner='',
                        notes=data['notes'] or '',
                    )
                    tx.last_edited_by = current_user.id
                    db.add(tx)
                    db.commit()
                    log = AuditLog(actor_id=current_user.id, action='create', tx_id=tx.id, details=f'[bulk upload] Created transaction for {tx.date}')
                    db.add(log)
                    db.commit()
                    created += 1
            except Exception as e:
                errors.append(f'Row {idx}: {e}')
                continue

        summary = dict(created=created, updated=updated, skipped=skipped, errors=errors)
        if errors:
            flash(f'Import completed with {len(errors)} issues. Created: {created}, Updated: {updated}, Skipped: {skipped}', 'warning')
        else:
            flash(f'Import completed. Created: {created}, Updated: {updated}, Skipped: {skipped}', 'success')
    return render_template('admin_upload.html', form=form, summary=summary)


@app.route('/admin/upload/template.csv')
@login_required
def admin_upload_template_csv():
    if not current_user.is_admin:
        abort(403)
    # CSV header aligned with the official Excel template headers
    headers = [
        'Date',
        'Number of Orders',
        'Total Voids',
        'Dine-In Cash',
        'Catering Orders (Exclude Biryani Trays)',
        'Biryani Trays',
        'Event Hall Rental & Food',
        'Paid to Employee (Add Name in Notes)',
        'Toast Card Sale',
        'Net Card Tips',
        'Doordash',
        'Uber Eats',
        'Grubhub',
        'Cancelled Orders',
        'Toast Fees',
        'Doordash Fees',
        'Uber Eats Fees',
        'Grubhub Fees',
        'Cash Verified By Owner',
        'Notes',
    ]
    import io, csv as _csv
    sio = io.StringIO()
    writer = _csv.writer(sio)
    writer.writerow(headers)
    csv_text = sio.getvalue()
    resp = make_response(csv_text)
    resp.headers['Content-Disposition'] = 'attachment; filename=BC_Sales_Template.csv'
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    return resp


@app.route('/admin/audit-logs')
@login_required
def audit_logs():
    if not current_user.is_admin:
        abort(403)
    db = SessionLocal()
    # filters
    page = int(request.args.get('page', 1))
    per_page = 50
    action = request.args.get('action')
    actor = request.args.get('actor')

    q = db.query(AuditLog)
    if action:
        q = q.filter(AuditLog.action == action)
    if actor:
        # try matching username
        actor_user = db.query(User).filter(User.username == actor).first()
        if actor_user:
            q = q.filter(AuditLog.actor_id == actor_user.id)
    total = q.count()
    logs = q.order_by(AuditLog.timestamp.desc()).offset((page-1)*per_page).limit(per_page).all()
    has_next = (page*per_page) < total
    has_prev = page > 1
    return render_template('audit_logs.html', logs=logs, page=page, has_next=has_next, has_prev=has_prev, action=action or '', actor=actor or '')


@app.cli.command('init-db')
def cli_init_db():
    """Initialize the database."""
    init_db()
    print('Initialized DB at', DB_PATH)


@app.cli.command('create-admin')
def cli_create_admin():
    """Create an admin user interactively."""
    init_db()
    db = SessionLocal()
    username = input('Admin username: ')
    password = input('Password: ')
    if db.query(User).filter_by(username=username).first():
        print('User exists')
        return
    admin = User.create(username=username, password=password, is_admin=True)
    db.add(admin)
    db.commit()
    print('Admin created')


# Temporary development helper: create or update an admin user from the browser.
# Only enabled when running in debug mode or when the request comes from localhost.
@app.route('/_dev/create_admin', methods=['GET', 'POST'])
def dev_create_admin():
    # restrict to local/dev environments
    remote = request.remote_addr or ''
    if not (app.debug or remote.startswith('127.') or remote == '::1'):
        abort(404)
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = (request.form.get('password') or '').strip()
        if not username or not password:
            return '<p>Missing username or password. <a href="">Back</a></p>'
        db = SessionLocal()
        existing = db.query(User).filter(User.username == username).first()
        if existing:
            existing.password_hash = generate_password_hash(password)
            existing.is_admin = True
            db.add(existing)
            db.commit()
            return f'<p>Updated admin <strong>{username}</strong>. <a href="{url_for("login")}">Login</a></p>'
        else:
            u = User.create(username=username, password=password, is_admin=True)
            db.add(u)
            db.commit()
            return f'<p>Created admin <strong>{username}</strong>. <a href="{url_for("login")}">Login</a></p>'
    # GET: simple form
    return ('<h2>Dev: create/update admin</h2>'
            '<form method="post">'
            'Username: <input name="username" /><br/>'
            'Password: <input type="password" name="password" /><br/>'
            '<button type="submit">Create / Update admin</button>'
            '</form>')







if __name__ == '__main__':
    init_db()
    try:
        if OPENPYXL_AVAILABLE:
            app.logger.info('openpyxl available: version %s', OPENPYXL_VERSION)
        else:
            app.logger.info('openpyxl NOT available: %s', globals().get('OPENPYXL_IMPORT_ERROR', 'unknown error'))
    except Exception:
        pass
    # prefer explicit port 50010 to avoid macOS services binding to 5000
    import sys
    port = 50010
    host = '127.0.0.1'
    # allow overriding via command-line --port or FLASK_RUN_PORT env
    for i, a in enumerate(sys.argv):
        if a == '--port' and i+1 < len(sys.argv):
            try:
                port = int(sys.argv[i+1])
            except Exception:
                pass
        if a == '--host' and i+1 < len(sys.argv):
            try:
                host = str(sys.argv[i+1])
            except Exception:
                pass
    port = int(os.environ.get('FLASK_RUN_PORT', port))
    host = os.environ.get('FLASK_RUN_HOST', host)
    # When running as a background/daemon (nohup), the Werkzeug reloader
    # may attempt to access stdin and fail with termios.error. Disable the
    # reloader to make background launches stable while keeping debug=True.
    app.run(debug=True, host=host, port=port, use_reloader=False)
